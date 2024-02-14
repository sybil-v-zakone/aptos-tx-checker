import json

from openpyxl import Workbook
import requests
from loguru import logger
from aptos_sdk.client import ApiError, ResourceNotFound

from config import USE_MOBILE_PROXY, CHANGE_IP_URL, TOKENS, USE_PROXY, RPC_URL
from constants import RESULT_JSON_PATH, COIN_STORE, Token, WALLET_ADDRESSES_PATH, PROXIES_PATH


def change_mobile_ip() -> None:
    try:
        if USE_MOBILE_PROXY:
            res = requests.get(CHANGE_IP_URL)

            if res.status_code == 200:
                logger.info("IP address changed successfully", send_to_tg=False)
            else:
                raise Exception("Failed to change IP address")

    except Exception as e:
        raise Exception(f"Encountered an error when changing ip address, check your proxy provider: {e}")


def read_from_txt(file_path: str) -> list[str]:
    try:
        with open(file_path, "r") as file:
            return [line.strip() for line in file]

    except Exception as e:
        raise Exception(f"Encountered an error while reading a txt file '{file_path}': {str(e)}")


def save_to_json(result: dict | str) -> None:
    try:
        if type(result) is dict:
            result = json.dumps(result, indent=4)

        with open(RESULT_JSON_PATH, 'w') as json_file:
            json_file.write(result)

    except Exception as e:
        raise Exception(f"Error while save json: {str(e)}")


def account_resource(account_address, resource_type, proxy):
    request = f"{RPC_URL}/v1/accounts/{account_address}/resource/{resource_type}"

    proxy = None if proxy is None else {"https": f"http://{proxy}"}

    response = requests.get(url=request, proxies=proxy)

    if response.status_code == 404:
        raise ResourceNotFound(resource_type, resource_type)

    if response.status_code >= 400:
        raise ApiError(f"{response.text} - {account_address}", response.status_code)

    return response.json()


def get_token_balance(token: Token, address, proxy) -> int | None:
    try:
        store_address = f"{COIN_STORE}<{token.contract_address}>"
        account_res = account_resource(address, store_address, proxy)
        value = int(account_res["data"]["coin"]["value"])

        return token.from_wei(value)

    except Exception as e:
        if token.contract_address in str(e):
            logger.error(f"{token.symbol} token not registered")
            return 0
        else:
            logger.error(f"Get coin data error: {str(e)}")


def get_addresses_and_proxies():
    addresses = read_from_txt(WALLET_ADDRESSES_PATH)
    proxies = read_from_txt(PROXIES_PATH)

    if USE_PROXY and len(proxies) != len(addresses):
        logger.error("Proxies and addresses should be one to one")
        exit()

    return addresses, proxies


def get_tokens_balance_batch(address: str, proxy: dict) -> dict:
    try:
        token_balances = {}

        for token in TOKENS:
            token_balances[token.symbol] = get_token_balance(token, address, proxy)

        return token_balances
    except Exception as e:
        logger.error(f"Exception in get token balance function: {str(e)}")


def save_to_exel(data: dict) -> None:
    wb = Workbook()
    ws = wb.active

    default_column_names = [
        'address',
        'pancakeswap',
        'liquidswap',
        'ditto',
        'tortuga',
        'sushiswap',
        'mercato',
        'merkle',
        'wapal',
        'topaz',
        'thalaswap',
        'bluemove',
        'amnis',
        'econia',
        'swapgpt',
        'kanalabs',
        'tx_count',
        'nft'
    ]

    token_column_names = []
    for token in TOKENS:
        token_column_names.append(token.symbol)

    total_column_names = [
        'quest_3_nft_count',
        'quest_2_nft_count',
        'quest_1_nft_count',
        'aptos_domains_count'
    ]

    column_names = default_column_names + token_column_names + total_column_names

    for col, name in enumerate(column_names, start=1):
        ws.cell(row=1, column=col, value=name)

    row = 2
    for address, address_data in data.items():
        ws.cell(row=row, column=1, value=address)
        for key, value in address_data.items():
            if isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    ws.cell(row=row, column=column_names.index(sub_key) + 1, value=sub_value)
            elif isinstance(value, list):
                value = ', '.join(value)
                ws.cell(row=row, column=column_names.index(key) + 1, value=value)
            else:
                ws.cell(row=row, column=column_names.index(key) + 1, value=value)
        row += 1

    wb.save("data/result.xlsx")
